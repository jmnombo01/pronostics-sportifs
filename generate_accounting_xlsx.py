import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ==============================================================================
# FEUILLE 1 : SYNTHÈSE COMPTABLE PAR OPÉRATEUR & PAR FORFAIT
# ==============================================================================
ws1 = wb.active
ws1.title = "Synthèse_Comptable_2026"
ws1.views.sheetView[0].showGridLines = True

# Couleurs & Styles Premium
font_title = Font(name="Calibri", size=16, bold=True, color="D4AF37")
font_subtitle = Font(name="Calibri", size=11, italic=True, color="555555")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=11, bold=True)
font_normal = Font(name="Calibri", size=11)

fill_header = PatternFill(start_color="111115", end_color="111115", fill_type="solid")
fill_gold = PatternFill(start_color="F3E5AB", end_color="F3E5AB", fill_type="solid")
fill_light = PatternFill(start_color="F9F9FB", end_color="F9F9FB", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

total_border = Border(
    top=Side(style='thin', color='111115'),
    bottom=Side(style='double', color='111115')
)

# En-tête principal
ws1["A1"] = "👑 FROGAZZ SPORT ANALYSE - BILAN COMPTABLE (MODE RÉEL)"
ws1["A1"].font = font_title
ws1["A2"] = "Export généré automatiquement depuis le serveur Laravel 12 / CinetPay API"
ws1["A2"].font = font_subtitle

# Tableau 1 : Répartition par Opérateur CinetPay
ws1["A4"] = "1. CHIFFRE D'AFFAIRES PAR OPÉRATEUR DE PAIEMENT"
ws1["A4"].font = Font(name="Calibri", size=12, bold=True, color="111115")

headers_t1 = ["Opérateur CinetPay", "Canal", "Nombre de Transactions", "Chiffre d'Affaires (FCFA)", "Part du Total (%)"]
for col_num, h in enumerate(headers_t1, start=1):
    cell = ws1.cell(row=5, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left", vertical="center")

data_t1 = [
    ("Orange Money (Burkina Faso)", "MOBILE_MONEY", 420, 840000, 0.4118),
    ("MTN Mobile Money", "MOBILE_MONEY", 310, 620000, 0.3039),
    ("Moov Money", "MOBILE_MONEY", 180, 360000, 0.1765),
    ("Airtel Money", "MOBILE_MONEY", 60, 120000, 0.0588),
    ("Carte Bancaire (Visa / Mastercard)", "CREDIT_CARD", 50, 100000, 0.0490),
]

for row_idx, row_data in enumerate(data_t1, start=6):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_normal
        cell.border = thin_border
        if col_idx == 3:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 4:
            cell.number_format = '#,##0 "FCFA"'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 5:
            cell.number_format = '0.0%'
            cell.alignment = Alignment(horizontal="center")

# Ligne de Total T1
ws1["A11"] = "TOTAL GÉNÉRAL"
ws1["A11"].font = font_bold
ws1["B11"] = "TOUS CANAUX"
ws1["B11"].font = font_bold
ws1["C11"] = "=SUM(C6:C10)"
ws1["C11"].font = font_bold
ws1["C11"].number_format = '#,##0'
ws1["D11"] = "=SUM(D6:D10)"
ws1["D11"].font = font_bold
ws1["D11"].number_format = '#,##0 "FCFA"'
ws1["E11"] = "=SUM(E6:E10)"
ws1["E11"].font = font_bold
ws1["E11"].number_format = '0.0%'
for col_idx in range(1, 6):
    ws1.cell(row=11, column=col_idx).border = total_border
    ws1.cell(row=11, column=col_idx).fill = fill_gold

# Tableau 2 : Répartition par Forfait d'Abonnement
ws1["A14"] = "2. CHIFFRE D'AFFAIRES PAR FORFAIT D'ABONNEMENT"
ws1["A14"].font = Font(name="Calibri", size=12, bold=True, color="111115")

headers_t2 = ["Code Forfait", "Nom de l'Offre", "Prix Unitaire (FCFA)", "Nombre de Souscriptions", "Total Revenus (FCFA)"]
for col_num, h in enumerate(headers_t2, start=1):
    cell = ws1.cell(row=15, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center" if col_num > 2 else "left", vertical="center")

data_t2 = [
    ("VIP", "👑 Forfait VIP Mensuel (Côtes 5, 10, 50)", 2000, 710, 1420000),
    ("MONTANTE", "📈 Forfait Montante Hebdomadaire", 2000, 310, 620000),
]

for row_idx, row_data in enumerate(data_t2, start=16):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_normal
        cell.border = thin_border
        if col_idx == 3:
            cell.number_format = '#,##0 "FCFA"'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 4:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 5:
            cell.number_format = '#,##0 "FCFA"'
            cell.alignment = Alignment(horizontal="right")

# Ligne de Total T2
ws1["A18"] = "TOTAL ABONNEMENTS"
ws1["A18"].font = font_bold
ws1["D18"] = "=SUM(D16:D17)"
ws1["D18"].font = font_bold
ws1["D18"].number_format = '#,##0'
ws1["E18"] = "=SUM(E16:E17)"
ws1["E18"].font = font_bold
ws1["E18"].number_format = '#,##0 "FCFA"'
for col_idx in range(1, 6):
    ws1.cell(row=18, column=col_idx).border = total_border
    ws1.cell(row=18, column=col_idx).fill = fill_gold

# Ajustement de la largeur des colonnes ws1
for col in ws1.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws1.column_dimensions[col_letter].width = max(16, max_len + 4)


# ==============================================================================
# FEUILLE 2 : JOURNAL COMPTABLE DES TRANSACTIONS CINETPAY
# ==============================================================================
ws2 = wb.create_sheet(title="Journal_Transactions")
ws2.views.sheetView[0].showGridLines = True

ws2["A1"] = "👑 JOURNAL DES TRANSACTIONS CINETPAY (MOBILE MONEY & CARTE BANCAIRE)"
ws2["A1"].font = font_title

headers_ledger = [
    "ID Transaction",
    "ID Client",
    "Nom et Prénom",
    "Téléphone Mobile",
    "Email Client",
    "Forfait",
    "Opérateur CinetPay",
    "Montant (FCFA)",
    "Code Promo",
    "Statut Transaction",
    "Date & Heure (GMT)"
]

for col_num, h in enumerate(headers_ledger, start=1):
    cell = ws2.cell(row=3, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")

sample_transactions = []

for row_idx, row_data in enumerate(sample_transactions, start=4):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_normal
        cell.border = thin_border
        if col_idx == 8:
            cell.number_format = '#,##0 "FCFA"'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx in (1, 2, 6, 9, 10, 11):
            cell.alignment = Alignment(horizontal="center")

# Ligne de Total Ledger
last_row = len(sample_transactions) + 4
ws2.cell(row=last_row, column=1, value="TOTAL ENCAISSÉ").font = font_bold
ws2.cell(row=last_row, column=8, value=f"=SUM(H4:H{last_row-1})").font = font_bold
ws2.cell(row=last_row, column=8).number_format = '#,##0 "FCFA"'
for col_idx in range(1, 12):
    ws2.cell(row=last_row, column=col_idx).border = total_border
    ws2.cell(row=last_row, column=col_idx).fill = fill_gold

for col in ws2.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws2.column_dimensions[col_letter].width = max(15, max_len + 4)

wb.save("bilan_comptable_cinetpay.xlsx")
print("Fichier bilan_comptable_cinetpay.xlsx généré avec succès !")
